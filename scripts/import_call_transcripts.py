"""
Import call transcripts from xlsx into Cosmos DB.

Usage:
    python scripts/import_call_transcripts.py

Requires:
    - azure-cosmos
    - azure-identity
    - openpyxl
"""
import asyncio
import logging
import sys
from azure.cosmos.aio import CosmosClient
from azure.identity.aio import AzureCliCredential

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
# Suppress Azure SDK verbose noise
for _name in ["azure", "azure.cosmos", "azure.identity", "azure.core"]:
    logging.getLogger(_name).setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# ── Config ──────────────────────────────────────────────────
COSMOS_ENDPOINT = "https://cosmos-2v3lfktkn4xam-gprag.documents.azure.com:443/"
DATABASE_NAME = "cosmos-db2v3lfktkn4xam-gprag"
CONTAINER_NAME = "call-transcripts"
XLSX_PATH = r"C:\Users\v-ktseng\AppData\Local\Temp\call_transcripts.xlsx"
BATCH_SIZE = 50


def parse_call_date(raw_date) -> str:
    """Parse date like '2026010912' → '2026-01-09'"""
    s = str(raw_date).strip()
    if len(s) >= 8:
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s


def load_xlsx(path: str) -> list[dict]:
    """Load xlsx and return list of document dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]  # ('客代', '通話ID', '通話日期', '推銷狀態', '通話文本')
    logger.info(f"Headers: {headers}")
    logger.info(f"Data rows: {len(rows) - 1}")

    documents = []
    for row in rows[1:]:
        customer_id = str(row[0]).strip() if row[0] else ""
        call_id = str(row[1]).strip() if row[1] else ""
        call_date = parse_call_date(row[2]) if row[2] else ""
        status = str(row[3]).strip() if row[3] else ""
        transcript = str(row[4]) if row[4] else ""

        doc = {
            "id": call_id,                     # unique document ID
            "customer_id": customer_id,         # partition key
            "call_id": call_id,
            "call_date": call_date,
            "status": status,                   # 成功 / 失敗
            "transcript": transcript,
            "source_file": "會員卡推廣名單通話文本_成功失敗各500人_去敏.xlsx",
        }
        documents.append(doc)

    wb.close()
    return documents


async def import_to_cosmos(documents: list[dict]):
    """Upsert documents into Cosmos DB."""
    credential = AzureCliCredential()

    async with CosmosClient(COSMOS_ENDPOINT, credential=credential) as client:
        db = client.get_database_client(DATABASE_NAME)
        container = db.get_container_client(CONTAINER_NAME)

        success = 0
        failed = 0

        for i in range(0, len(documents), BATCH_SIZE):
            batch = documents[i:i + BATCH_SIZE]
            for doc in batch:
                try:
                    await container.upsert_item(doc)
                    success += 1
                except Exception as e:
                    failed += 1
                    logger.error(f"Failed to upsert {doc['id']}: {e}")

            logger.info(f"Progress: {min(i + BATCH_SIZE, len(documents))}/{len(documents)} "
                        f"(success={success}, failed={failed})")

    logger.info(f"✅ Import complete: {success} succeeded, {failed} failed")


async def main():
    logger.info(f"Loading xlsx from: {XLSX_PATH}")
    documents = load_xlsx(XLSX_PATH)

    logger.info(f"Importing {len(documents)} documents to Cosmos DB...")
    logger.info(f"  Endpoint: {COSMOS_ENDPOINT}")
    logger.info(f"  Database: {DATABASE_NAME}")
    logger.info(f"  Container: {CONTAINER_NAME}")

    await import_to_cosmos(documents)


if __name__ == "__main__":
    asyncio.run(main())
